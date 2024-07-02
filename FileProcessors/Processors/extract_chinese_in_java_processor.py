import os.path
import re

from fuzzywuzzy import fuzz
from openpyxl.reader.excel import load_workbook

from FileProcessors.abstract_file_processor import AbstractFileProcessor


class ExtractChineseInJavaProcessor(AbstractFileProcessor):

    def __init__(self):
        super().__init__()
        self.generators = list()
        self.template_excel = r"E:\temp\俄语版.xlsx"
        self.excel_sheets = self._init_sheets()

    def get_name(self):
        return "extractChineseInJavaProcessor"

    def support_process(self, file_path):
        return file_path.endswith('.java')

    def process(self, file_path):
        java_package = None
        file_chinese_list = list()
        with open(file_path, "r", encoding="utf-8") as f:
            multi_annotation = False
            row_num = 0
            for line in f:
                row_num += 1
                line = line.strip()
                if line.startswith("package"):
                    java_package = line.strip("package").strip(";")
                    file_name = os.path.basename(file_path)
                    java_package = java_package + "(" + file_name + ")"
                # 不处理注解的中文
                if line.startswith("@"):
                    continue
                # 单行注释的处理
                if line.startswith("//"):
                    continue

                # 多行注释的处理
                if line.startswith("/*"):
                    multi_annotation = True
                if line.startswith("*/"):
                    multi_annotation = False
                    continue
                if multi_annotation:
                    continue
                if line.find("@Api") != -1:
                    continue
                # 行内的处理
                current_chinese = self._process_single_line(line)
                if len(current_chinese) > 0 and self._contains_chinese(current_chinese):
                    # 判断本行是否不包含中文，不包含中文的就去掉
                    chinese_include_row = list()
                    row_number = "行号:" + str(row_num)
                    chinese_include_row.append(row_number)
                    chinese_include_row.append(line)
                    chinese_include_row.append(current_chinese)
                    # 查找是否有匹配的,模糊匹配
                    fuzzy_match_cn, ru = self._fuzzy_match_from_excel(current_chinese)

                    chinese_include_row.append(fuzzy_match_cn)
                    chinese_include_row.append(ru)
                    file_chinese_list.append(chinese_include_row)
        if len(file_chinese_list) > 0:
            for generator in self.generators:
                generator.generate(java_package, file_chinese_list)

    def post_process(self):
        for generator in self.generators:
            generator.post_process()

    @staticmethod
    def _process_single_line(line):
        into_chinese = False
        current_chinese = ""
        strip_annotation_index = line.find("//")
        if strip_annotation_index != -1:
            line = line[:strip_annotation_index]
        for content in line:
            if content == "\"":
                if into_chinese:
                    into_chinese = False
                else:
                    into_chinese = True
                    continue
            if into_chinese:
                current_chinese += content
        return current_chinese

    @staticmethod
    def _contains_chinese(text):
        pattern = re.compile(r'[\u4e00-\u9fff]')  # 匹配中文字符的正则表达式范围
        return bool(pattern.search(text))

    def _fuzzy_match_from_excel(self, current_chinese):
        max_ratios = dict()
        for sheet in self.excel_sheets:
            for row in sheet.iter_rows(values_only=True):
                if row:
                    compare_cell = row[1]
                    if fuzz.ratio(compare_cell, current_chinese.strip("\n")) >= 90 and fuzz.ratio(
                            current_chinese.strip("\n"), compare_cell) >= 90:
                        current_max_ratio = max(fuzz.ratio(compare_cell, current_chinese.strip("\n")),
                                                fuzz.ratio(current_chinese.strip("\n"), compare_cell))
                        if len(max_ratios) == 0:
                            max_ratios["cell"] = row
                            max_ratios["ratio"] = current_max_ratio
                        else:
                            if current_max_ratio >= max_ratios["ratio"]:
                                if current_max_ratio == max_ratios["ratio"] and row[2] != current_chinese.strip("\n"):
                                    # 匹配度相等条件下看第二个中文是否完全相等，第二个完全相等就设置，否则不设置
                                    continue
                                else:
                                    max_ratios["cell"] = row
                                    max_ratios["ratio"] = current_max_ratio

        if len(max_ratios) == 0:
            return "", ""
        else:
            cell = max_ratios["cell"]
            return cell[1], cell[2] if cell[2] is not None else ""

    def _init_sheets(self):
        workbook = load_workbook(self.template_excel)
        sheets = list()
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheets.append(sheet)
        return sheets
