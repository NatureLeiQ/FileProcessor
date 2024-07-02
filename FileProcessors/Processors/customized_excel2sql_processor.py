from openpyxl import load_workbook

from FileProcessors.abstract_file_processor import AbstractFileProcessor
from Utils.file_utils import resolve_file_path


class CustomizedExcel2SQLProcessor(AbstractFileProcessor):
    def __init__(self):
        super().__init__()
        self.excel_sheets = self._init_sheets()

    def support_process(self, file_path):
        return file_path.endswith('.sql')

    def process(self, file_path):
        contents = ""
        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                strip_annotation_index = line.find("VALUES")
                if strip_annotation_index == -1:
                    continue
                else:
                    process_line = line[strip_annotation_index:]  # value部分的内容
                process_line = process_line.replace("VALUES", "").replace("'", "").strip("(").strip(")")
                process_line_list = process_line.split(", ")
                code = process_line_list[2]
                zh = process_line_list[3]
                match_ru = self._get_match_translate(code, zh)
                if match_ru != "":
                    contents = contents + self._concat_sql(match_ru, code, zh) + "\n"
            if contents != "":
                for generator in self.generators:
                    resolved_file_path = resolve_file_path(file_path)
                    russia_properties_name = resolved_file_path["file_name"] + "_config_ru"
                    generator.generate(russia_properties_name, contents)

    def post_process(self):
        pass

    def get_name(self):
        return "customizedExcel2SQLProcessor"

    @staticmethod
    def _init_sheets():
        excel_file_path = r"H:\PythonProjects\i18n\俄语版.xlsx"
        if excel_file_path is None:
            return
        workbook = load_workbook(excel_file_path)
        sheets = list()

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheets.append(sheet)
        return sheets

    def _get_match_translate(self, match_code, match_zh):

        for sheet in self.excel_sheets:
            for row in sheet.iter_rows(values_only=True):
                if row:
                    code_cell = row[0]
                    zh_cell = row[1]
                    if zh_cell is not None:
                        zh_cell = zh_cell.strip()
                    if match_zh is not None:
                        match_zh = match_zh.strip()
                    if code_cell == match_code and zh_cell == match_zh:
                        return row[2] if row[2] is not None else ""
        return ""

    @staticmethod
    def _concat_sql(set_value, code, zh):
        sql = "UPDATE i18n SET ru = '" + set_value + "' WHERE `code` = '" + code + "' AND zh = '" + zh + "';"
        return sql



