from openpyxl import load_workbook

from FileProcessors.abstract_file_processor import AbstractFileProcessor
from Utils.file_utils import resolve_file_path


class CustomizedExcel2PropertiesProcessor(AbstractFileProcessor):
    def __init__(self):
        super().__init__()
        self.excel_sheets = self._init_sheets()

    def support_process(self, file_path):
        return file_path.endswith('.properties')

    def process(self, file_path):
        contents = ""
        with open(file_path, "r", encoding="utf-8") as file:

            for line in file:
                line = line.strip()
                strip_annotation_index = line.find("=")
                if strip_annotation_index == -1:
                    continue
                else:
                    en_line = line[:strip_annotation_index]  # 英文内容
                contents = contents + en_line + "=" + self._get_match_translate(en_line) + "\n"
            for generator in self.generators:
                resolved_file_path = resolve_file_path(file_path)
                russia_properties_name = resolved_file_path["file_name"]+"_ru_RU"
                generator.generate(russia_properties_name, contents)

    def post_process(self):
        pass

    def get_name(self):
        return "customizedExcel2PropertiesProcessor"

    @staticmethod
    def _init_sheets():
        excel_file_path = ""
        if excel_file_path is None or excel_file_path == "":
            return
        workbook = load_workbook(excel_file_path)
        sheets = list()

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheets.append(sheet)
        return sheets

    def _get_match_translate(self, match):
        for sheet in self.excel_sheets:
            for row in sheet.iter_rows(values_only=True):
                if row:
                    compare_cell = row[0]
                    if compare_cell == match:
                        return row[2] if row[2] is not None else ""
        return ""
