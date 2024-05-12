import xlrd
from openpyxl import load_workbook

from FileTraverse.TraverseStrategies.Strategies.FileContentStrategy.abstract_file_scanner import AbstractFileScanner
from FileTraverse.TraverseStrategies.Strategies.FileContentStrategy.utils.scanner_text_utils import content_match
from Utils.file_mime_type_enums import FileMimeTypeEnums


class ExcelScanner(AbstractFileScanner):

    def __init__(self):
        super().__init__()

    def support_scan(self, file_info):
        mime_type_enum = file_info["mime_type_enum"]
        match mime_type_enum:
            case FileMimeTypeEnums.xls | FileMimeTypeEnums.xlsx:
                return True
        return False

    def get_name(self):
        return "excelScanner"

    def scan(self, file_info):
        mime_type_enum = file_info["mime_type_enum"]
        match mime_type_enum:
            case FileMimeTypeEnums.xls:
                return self._scan_xls(file_info)
            case FileMimeTypeEnums.xlsx:
                return self._scan_xlsx(file_info)
        return False

    def _scan_xlsx(self, file_info):
        file_path = file_info["file_path"]
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # 遍历工作表的每一行和每一列
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and content_match(str(cell.value), self.match_text, self.fuzzy_match):
                        return True
        return False

    def _scan_xls(self, file_info):
        file_path = file_info["file_path"]
        workbook = xlrd.open_workbook(file_path)
        for sheet in workbook.sheets():
            for row in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col_idx)
                    if cell_value is not None and content_match(str(cell_value), self.match_text, self.fuzzy_match):
                        return True
        return False
